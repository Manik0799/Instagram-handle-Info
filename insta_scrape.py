import requests
from bs4 import BeautifulSoup
import json
import openpyxl


def create_excel_sheet(data_list):
    # Creating a new workbook
    book = openpyxl.Workbook()
    sheet = book.active

    # Creating the headers for the excel sheet
    headers = ["Username", "Full Name", "Profile Picture", "Verified",
               "Private", "Bio Text", "External Links", "Followers", "Following"]

    i = 1
    for header in headers:
        sheet.cell(row=1, column=i).value = header
        i += 1

    # Writing the fetched data
    j = 3

    for user in data_list:
        i = 1
        for entry in user:
            sheet.cell(row=j, column=i).value = entry
            i += 1
        j += 1

    # Saving the workbook
    book.save("Insta_info.xlsx")


def insta_bio(users):

    data_list = []

    for user in users:
        url = f"https://www.instagram.com/{user}/"
        response = requests.get(url)

        if response.status_code == 200:
            # Parsing HTML
            soup = BeautifulSoup(response.text, 'html.parser')

            scripts = soup.find_all("script")
            # 4th script tag out of the 16 tags in total are of use
            bio_script = scripts[4]

            main_content = bio_script.contents[0]
            # Using -1 to remove the semi colon at the end
            data = main_content[main_content.find('{"config"'): -1]

            bio_json = json.loads(data)

            bio_json = bio_json["entry_data"]["ProfilePage"][0]["graphql"]["user"]

            # Appending the data_list
            data_list.append([
                bio_json["username"],
                bio_json["full_name"],
                bio_json["profile_pic_url_hd"],
                bio_json["is_verified"],
                bio_json["is_private"],
                bio_json["biography"],
                bio_json["external_url"],
                bio_json["edge_followed_by"]["count"],
                bio_json["edge_follow"]["count"]
            ])

        else:
            data_list.append([f"{user} not found"]*9)

    create_excel_sheet(data_list)


if __name__ == "__main__":

    users = ["manik_0799", "yestheory", "leomessi", "willsmith"]
    print("FETCHING ... ")
    insta_bio(users)
    print("DATA SUCCESSFULLY WRITTEN...")
